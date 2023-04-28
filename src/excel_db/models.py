import typing

from openpyxl.cell import Cell
from returns import returns


class ExcelModel:
    columns: list['Column']

    @classmethod
    def as_table(cls, **kwargs) -> 'ExcelTableDefinition':
        return ExcelTableDefinition(cls, **kwargs)

    def __init__(
            self,
            table: 'ExcelTable',
            idx: int,
            row_num: int,
    ):
        self.table = table
        self.idx = idx
        self.row_num = row_num

    def __eq__(self, other: 'ExcelModel') -> bool:
        if other is None or not isinstance(other, ExcelModel):
            return False

        return (
                self.table == other.table
                and self.idx == other.idx
                and self.row_num == other.row_num
        )

    @returns(dict)
    def as_dict(self) -> dict[str, typing.Any]:
        for column in self.columns:
            yield column.name, column.__get__(self)

    def set_dict(
            self,
            mapping: typing.Mapping[str, typing.Any] = None,
            /,
            **kwargs,
    ):
        if mapping is not None:
            for k, v in mapping.items():
                setattr(self, k, v)
        for k, v in kwargs.items():
            setattr(self, k, v)

    def cell(self, col_num: int) -> Cell:
        return self.table.cell(self.row_num, col_num)

    def cell0(self, col_idx: int) -> Cell:
        return self.cell(col_idx + 1)

    def cella(self, attr: str) -> Cell:
        return self.cell(getattr(self.table, attr).col_num)


from .columns import Column  # noqa: E402
from .tables import ExcelTableDefinition, ExcelTable  # noqa: E402
