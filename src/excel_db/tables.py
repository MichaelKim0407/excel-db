import typing
from functools import cached_property

from openpyxl.worksheet.worksheet import Worksheet

from .exceptions import ColumnNotFound
from .utils.descriptors import BasePropertyDescriptor


class ExcelTableDefinition(BasePropertyDescriptor['ExcelDB']):
    title_row: int = 1

    def __init__(
            self,
            model: typing.Type['ExcelModel'],
            **kwargs,
    ):
        super().__init__(**kwargs)
        self.model = model

    def _add_to_class(self):
        if not hasattr(self.obj_type, 'tables'):
            self.obj_type.tables = []
        self.obj_type.tables.append(self)

    @cached_property
    def _cache_name(self) -> str:
        return f'_ws_{self.attr}'

    _f_initialize = None

    def _initialize_title_row(self, db: 'ExcelDB', ws: Worksheet):
        for i, column in enumerate(self.model.columns):
            ws.cell(self.title_row, i + 1, column.name)

    def _initialize_default(self, db: 'ExcelDB', ws: Worksheet):
        pass

    @property
    def _initialize_method(self):
        if self._f_initialize is None:
            return self._initialize_default
        else:
            return self._f_initialize

    def initialize(self, db: 'ExcelDB', ws: Worksheet):
        self._initialize_title_row(db, ws)
        self._initialize_method(db, ws)  # noqa: pycharm

    def initializer(self, f_initialize):
        self._f_initialize = f_initialize
        return self

    def _get_default(self, db: 'ExcelDB') -> Worksheet:
        if self.name in db.wb:
            return db.wb[self.name]
        else:
            ws = db.wb.create_sheet(self.name)
            self.initialize(db, ws)
            return ws

    def _get(self, db: 'ExcelDB') -> 'ExcelTable':
        if self.attr not in db.__dict__:
            ws = self._get_method(db)
            db.__dict__[self._cache_name] = ws
        ws = db.__dict__[self._cache_name]
        return ExcelTable(db, self, ws)

    def _set_default(self, db: 'ExcelDB', ws: Worksheet) -> Worksheet:
        if self.name in db.wb:
            del db.wb[self.name]
        copy: Worksheet = db.wb.copy_worksheet(ws)
        copy.title = self.name
        return copy

    def _set(self, db: 'ExcelDB', ws: typing.Union[Worksheet, 'ExcelTable']):
        if isinstance(ws, ExcelTable):
            ws = ws.ws
        copy = self._set_method(db, ws)
        db.__dict__[self._cache_name] = copy

    def _delete_default(self, db: 'ExcelDB'):
        del db.wb[self.name]

    def _delete(self, db: 'ExcelDB'):
        if self.attr in db.__dict__:
            del db.__dict__[self._cache_name]
        self._delete_method(db)


class ExcelTable:
    def __init__(
            self,
            db: 'ExcelDB',
            table: ExcelTableDefinition,
            ws: Worksheet,
    ):
        self.db = db
        self.table = table
        self.ws = ws

        self._columns_cache = {}

    def __eq__(self, other: 'ExcelTable') -> bool:
        if other is None or not isinstance(other, ExcelTable):
            return False

        return (
                self.db == other.db
                and self.table == other.table
                and self.ws == other.ws
        )

    @property
    def model(self) -> typing.Type['ExcelModel']:
        return self.table.model

    def _get_row_num(self, idx: int) -> int:
        return self.table.title_row + idx + 1

    def _get_range(
            self,
            s: slice = None,
            *,
            start=None,
            stop=None,
            step=None,
    ) -> list[int]:
        if s is not None:
            start = s.start
            stop = s.stop
            step = s.step

        return list(range(self.ws.max_row - self.table.title_row))[start:stop:step]

    def __getitem__(self, idx: int | slice) -> typing.Union['ExcelModel', list['ExcelModel']]:
        if isinstance(idx, slice):
            return [
                self[i]
                for i in self._get_range(idx)
            ]
        return self.model(self, idx, self._get_row_num(idx))

    def __len__(self) -> int:
        return len(self._get_range())

    def __iter__(self) -> typing.Iterator['ExcelModel']:
        for i in self._get_range():
            yield self[i]

    def _get_column_def(self, attr: str) -> 'Column':
        for column in self.model.columns:
            if column.attr == attr:
                return column
        raise AttributeError(attr)

    def _get_col_num(self, name: str) -> int:
        for cell in self.ws[self.table.title_row]:
            if cell.value == name:
                return cell.column
        raise ColumnNotFound(name)

    def _get_column(self, attr: str) -> 'ExcelColumn':
        if attr not in self._columns_cache:
            column_def = self._get_column_def(attr)
            col_num = self._get_col_num(column_def.name)
            self._columns_cache[attr] = ExcelColumn(self, column_def, col_num)
        return self._columns_cache[attr]

    def __getattr__(self, attr: str) -> 'ExcelColumn':
        return self._get_column(attr)


from .db import ExcelDB  # noqa: E402
from .models import ExcelModel  # noqa: E402
from .columns import Column, ExcelColumn  # noqa: E402
