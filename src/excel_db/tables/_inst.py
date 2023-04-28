import typing

from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

from ._def import TTableDef
from ..db import TDB
from ..exceptions import ColumnNotFound
from ..models import TModel


class ExcelTable(typing.Generic[TDB, TModel, TTableDef]):
    def __init__(
            self,
            db: TDB,
            table: TTableDef,
            ws: Worksheet,
    ):
        self.db = db
        self.table = table
        self.ws = ws

        self._columns_cache = {}

    def __eq__(self, other: typing.Self) -> bool:
        if other is None or not isinstance(other, ExcelTable):
            return False

        return (
                self.db == other.db
                and self.table == other.table
                and self.ws == other.ws
        )

    @property
    def model(self) -> typing.Type[TModel]:
        return self.table.model

    def get_row_num(self, idx: int) -> int:
        return self.table.title_row + idx + 1

    def get_idx(self, row_num: int) -> int:
        return row_num - self.table.title_row - 1

    def _get_range(
            self,
            s: slice = None,
            *,
            start=None,
            stop=None,
            step=None,
    ) -> list[int]:
        if s is not None:
            start = s.start
            stop = s.stop
            step = s.step

        return list(range(self.ws.max_row - self.table.title_row))[start:stop:step]

    def __getitem__(self, idx: int | slice) -> typing.Union[TModel, list[TModel]]:
        if isinstance(idx, slice):
            return [
                self[i]
                for i in self._get_range(idx)
            ]
        return self.model(self, idx, self.get_row_num(idx))

    def __len__(self) -> int:
        return len(self._get_range())

    def __iter__(self) -> typing.Iterator[TModel]:
        for i in self._get_range():
            yield self[i]

    def _get_column_def(self, attr: str) -> 'TColumnDef':
        for column in self.model.columns:
            if column.attr == attr:
                return column
        raise AttributeError(attr)

    def _get_col_num(self, name: str) -> int:
        for cell in self.ws[self.table.title_row]:
            if cell.value == name:
                return cell.column
        raise ColumnNotFound(name)

    def _get_column(self, attr: str) -> 'TColumn':
        if attr not in self._columns_cache:
            column_def = self._get_column_def(attr)
            col_num = self._get_col_num(column_def.name)
            from ..columns import ExcelColumn
            self._columns_cache[attr] = ExcelColumn(self, column_def, col_num)
        return self._columns_cache[attr]

    def __getattr__(self, attr: str) -> 'TColumn':
        return self._get_column(attr)

    def cell(self, row_num, col_num) -> Cell:
        return self.ws.cell(row_num, col_num)

    def new(self) -> TModel:
        self.ws.append([])
        row_num = self.ws._current_row  # noqa: pycharm
        return self[self.get_idx(row_num)]


TTable = typing.TypeVar('TTable', bound=ExcelTable)

if typing.TYPE_CHECKING:
    from ..columns import TColumnDef, TColumn