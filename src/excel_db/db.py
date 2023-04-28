import os
import typing
from functools import cached_property

from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook

from .utils.class_collector import CollectorMeta, ListCollector


class ExcelDB(metaclass=CollectorMeta):
    tables: ListCollector['TTableDef']

    MODE_OPEN = 'o'  # open an existing file; FileNotFoundError
    MODE_CREATE = 'n'  # create a new file; FileExistsError
    MODE_OVERWRITE = 'w'  # overwrite and existing file; FileNotFoundError
    MODE_OPEN_CREATE = 'a'  # open an existing file; create a new file
    MODE_CREATE_OVERWRITE = 'x'  # create a new file; overwrite an existing file

    def __init__(
            self,
            filename: str,
            *,
            mode: str = MODE_OPEN_CREATE,
    ):
        self.filename = filename
        self.mode = mode

        self.ws_cache = {}

    @cached_property
    def wb(self) -> Workbook:
        if os.path.exists(self.filename):
            if self.mode in (self.MODE_OPEN, self.MODE_OPEN_CREATE):
                return load_workbook(self.filename)
            elif self.mode in (self.MODE_OVERWRITE, self.MODE_CREATE_OVERWRITE):
                return Workbook()
            else:
                raise FileExistsError(self.filename)
        else:
            if self.mode in (self.MODE_CREATE, self.MODE_OPEN_CREATE, self.MODE_CREATE_OVERWRITE):
                return Workbook()
            else:
                raise FileNotFoundError(self.filename)

    def save_as(self, filename: str):
        self.wb.save(filename)

    def save(self):
        self.save_as(self.filename)


TDB = typing.TypeVar('TDB', bound=ExcelDB)

if typing.TYPE_CHECKING:
    from .tables import TTableDef
