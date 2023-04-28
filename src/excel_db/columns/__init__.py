import typing
from functools import cached_property

from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter

from ..utils.descriptors import BasePropertyDescriptor


class Column(BasePropertyDescriptor['ExcelModel']):
    cache: bool = True

    def _add_to_class(self):
        if not hasattr(self.obj_type, 'columns'):
            self.obj_type.columns = []
        self.obj_type.columns.append(self)

    @cached_property
    def _cache_name(self) -> str:
        return f'_v_{self.attr}'

    def _get_col_num(self, row: 'ExcelModel') -> int:
        return getattr(row.table, self.attr).col_num

    def _get_cell(self, row: 'ExcelModel') -> Cell:
        return row.table.ws.cell(row.row_num, self._get_col_num(row))

    def _to_python(self, value):
        return value

    def _get_default(self, row: 'ExcelModel', cell: Cell):
        return self._to_python(cell.value)

    validators = ()

    def _validate(self, row: 'ExcelModel', value, cell: Cell):
        for validator in self.validators:
            validator(row, value, cell)

    def validator(self, f_validate):
        if isinstance(self.validators, tuple):
            self.validators = list(self.validators)
        self.validators.append(f_validate)
        return self

    _f_handle_error = None

    def _handle_error_default(self, row: 'ExcelModel', cell: Cell, ex: Exception):
        raise

    @property
    def _handle_error_method(self):
        if self._f_handle_error is None:
            return self._handle_error_default
        else:
            return self._f_handle_error

    def _handle_error(self, row: 'ExcelModel', cell: Cell, ex: Exception):
        return self._handle_error_method(row, cell, ex)  # noqa: pycharm

    def error_handler(self, f_handle_error):
        self._f_handle_error = f_handle_error
        return self

    def _get_nocache(self, row: 'ExcelModel'):
        cell = self._get_cell(row)
        try:
            value = self._get_method(row, cell)
            self._validate(row, value, cell)
            return value
        except Exception as ex:
            return self._handle_error(row, cell, ex)

    def _get(self, row: 'ExcelModel'):
        if self.cache:
            if self._cache_name not in row.__dict__:
                value = self._get_nocache(row)
                row.__dict__[self._cache_name] = value
            return row.__dict__[self._cache_name]
        else:
            return self._get_nocache(row)

    def _from_python(self, value):
        return value

    def _set_default(self, row: 'ExcelModel', value, cell: Cell):
        cell.value = self._from_python(value)

    def _set(self, row: 'ExcelModel', value):
        cell = self._get_cell(row)
        self._validate(row, value, cell)
        self._set_method(row, value, cell)
        if self.cache:
            row.__dict__[self._cache_name] = value

    def _delete_default(self, row: 'ExcelModel', cell: Cell):
        cell.value = None

    def _delete(self, row: 'ExcelModel'):
        self._delete_method(row, self._get_cell(row))
        if self.cache:
            if self._cache_name in row.__dict__:
                del row.__dict__[self._cache_name]


class ExcelColumn:
    def __init__(
            self,
            table: 'ExcelTable',
            column_def: Column,
            col_num: int,
    ):
        self.table = table
        self.column_def = column_def
        self.col_num = col_num

    def __eq__(self, other: 'ExcelColumn') -> bool:
        if other is None or not isinstance(other, ExcelColumn):
            return False

        return (
                self.table == other.table
                and self.column_def == other.column_def
                and self.col_num == other.col_num
        )

    def __getitem__(self, idx: int | slice):
        if isinstance(idx, slice):
            return [
                self.column_def.__get__(row)
                for row in self.table[idx]
            ]

        return self.column_def.__get__(self.table[idx])

    def __iter__(self) -> typing.Iterator:
        for row in self.table:
            yield self.column_def.__get__(row)

    def __setitem__(self, idx: int | slice, value):
        if isinstance(idx, slice):
            for row, v in zip(self.table[idx], value, strict=True):
                self.column_def.__set__(row, v)
            return

        self.column_def.__set__(self.table[idx], value)

    def __delitem__(self, idx: int | slice):
        if isinstance(idx, slice):
            for row in self.table[idx]:
                self.column_def.__delete__(row)
            return

        self.column_def.__delete__(self.table[idx])

    def cell(self, row_num: int) -> Cell:
        return self.table.cell(row_num, self.col_num)

    def cell0(self, idx: int) -> Cell:
        return self.cell(self.table.get_row_num(idx))

    @property
    def cells(self) -> typing.Sequence[Cell]:
        col_letter = get_column_letter(self.col_num)
        return self.table.ws[col_letter][self.table.get_row_num(-1):]


from ..tables import ExcelTable  # noqa: E402
from ..models import ExcelModel  # noqa: E402
