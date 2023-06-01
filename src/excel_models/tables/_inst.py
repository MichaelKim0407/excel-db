import typing

from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from excel_models.exceptions import ColumnNotFound, DuplicateColumn
from excel_models.typing import AbstractTable, TDB, TModel, TTableDef, TColumn


class ExcelTable(AbstractTable):
    def __init__(
            self,
            db: TDB,
            table_def: TTableDef,
            ws: Worksheet,
    ):
        self.db = db
        self.table_def = table_def
        self.ws = ws

        self.columns_cache = {}
        self.not_found = {}
        self.not_defined = []

    @property
    def model(self) -> typing.Type[TModel]:
        return self.table_def.model

    @property
    def title_row(self) -> int:
        return self.table_def.title_row

    @property
    def columns(self) -> typing.Sequence[TColumn]:
        return tuple(self.columns_cache.values())

    def _clear_cache(self):
        self.columns_cache.clear()
        self.not_found.clear()
        self.not_defined.clear()

    def find_columns(self):
        self._clear_cache()

        for cell in self.ws[self.title_row]:
            if cell.value is None:
                continue

            defined = False
            for column_def in self.model.column_defs:
                column = column_def.match_column(self, cell.column)
                if column is None:
                    continue
                self.columns_cache[column_def.attr] = column
                defined = True
                # There may be multiple column definitions to the same Excel column, so we keep going.

            if not defined:
                self.not_defined.append(cell.value)

        for column_def in self.model.column_defs:
            if column_def.attr in self.columns_cache:
                continue
            self.not_found[column_def.attr] = column_def

        self._check_found_columns()

    def _check_found_columns(self):
        col_nums = set()
        for column in self.columns:
            for n in column.occupied_col_nums:
                if n in col_nums:
                    raise DuplicateColumn(f'{column.column_def.name} on {n}')
                col_nums.add(n)

    def init_columns(self):
        self._clear_cache()

        col_num = 1
        for column_def in self.model.column_defs:
            column, width = column_def.init_column(self, col_num)
            self.columns_cache[column_def.attr] = column
            col_num += width

    def get_title(self, col_num: int) -> str:
        return self.ws.cell(self.title_row, col_num).value

    def set_title(self, col_num: int, name: str) -> None:
        self.ws.cell(self.title_row, col_num, name)

    def __eq__(self, other: typing.Self) -> bool:
        if other is None or not isinstance(other, ExcelTable):
            return False

        return (
                self.db == other.db
                and self.table_def == other.table_def
                and self.ws == other.ws
        )

    def __len__(self) -> int:
        return self.max_row - self.title_row

    def get_row_num(self, idx: int) -> int:
        return self.title_row + idx + 1

    def get_idx(self, row_num: int) -> int:
        return row_num - self.title_row - 1

    def _get_range(
            self,
            s: slice = None,
            *,
            start=None,
            stop=None,
            step=None,
    ) -> list[int]:
        if s is not None:
            start = s.start
            stop = s.stop
            step = s.step

        return list(range(len(self)))[start:stop:step]

    def __getitem__(self, idx: int | slice) -> TModel | list[TModel]:
        if isinstance(idx, slice):
            return [
                self[i]
                for i in self._get_range(idx)
            ]
        return self.model(self, idx, self.get_row_num(idx))

    def __iter__(self) -> typing.Iterator[TModel]:
        for i in self._get_range():
            yield self[i]

    def __getattr__(self, attr: str) -> TColumn:
        if attr in self.not_found:
            raise ColumnNotFound(self.not_found[attr].name)
        elif attr in self.columns_cache:
            return self.columns_cache[attr]
        else:
            raise AttributeError(attr)

    def new(self) -> TModel:
        self.ws.append([])
        row_num = self.ws._current_row  # noqa: pycharm
        return self[self.get_idx(row_num)]

    @property
    def max_col(self) -> int:
        return self.ws.max_column

    @property
    def max_column_letter(self) -> str:
        return get_column_letter(self.max_col)

    @property
    def max_row(self) -> int:
        return self.ws.max_row

    @property
    def data_rows(self) -> typing.Iterable[int]:
        return range(self.title_row + 1, self.max_row + 1)

    def cell(self, row_num: int, col_num: int) -> Cell:
        return self.ws.cell(row_num, col_num)

    def row(self, row_num: int) -> typing.Sequence[Cell]:
        return next(self.ws.iter_rows(min_row=row_num, max_row=row_num))

    def col(self, col_num: int, *, data_only: bool = False) -> typing.Sequence[Cell]:
        if data_only:
            min_row = self.title_row + 1
        else:
            min_row = None
        return next(self.ws.iter_cols(min_col=col_num, max_col=col_num, min_row=min_row))

    @property
    def _max_notnone_col(self) -> int:
        col = self.max_col
        while col > 0:
            for cell in self.col(col):
                if cell.value is not None:
                    return col
            col -= 1
        return col

    def trim_cols(self) -> int:
        max_col = self.max_col
        col = self._max_notnone_col
        if col == max_col:
            return col
        self.ws.delete_cols(col + 1, max_col - col)
        return col

    @property
    def _max_notnone_row(self) -> int:
        row = self.max_row
        while row > 0:
            for cell in self.row(row):
                if cell.value is not None:
                    return row
            row -= 1
        return row

    def trim_rows(self) -> int:
        max_row = self.max_row
        row = self._max_notnone_row
        if row == max_row:
            return row
        self.ws.delete_rows(row + 1, max_row - row)
        return row

    def trim(self) -> tuple[int, int]:
        return self.trim_cols(), self.trim_rows()

    @property
    def _filter_ref_str(self) -> str:
        return f'A{self.title_row}:{self.max_column_letter}{self.max_row}'

    def add_filter(self) -> None:
        self.ws.auto_filter.ref = self._filter_ref_str
