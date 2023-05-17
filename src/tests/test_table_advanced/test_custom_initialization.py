from openpyxl.worksheet.worksheet import Worksheet

from excel_db.columns import Column
from excel_db.db import ExcelDB
from excel_db.models import ExcelModel


class User(ExcelModel):
    id = Column()
    name = Column()


class MyDB(ExcelDB):
    users = User.as_table(title_row=2)

    @users.initializer
    def users(self, ws: Worksheet):
        ws.cell(1, 1, 'hello world')


def test_custom_init(tmp_path):
    tmp_excel_file = str(tmp_path / 'db.xlsx')
    db = MyDB(tmp_excel_file)
    _ = db.users
    assert db.wb['users'].cell(1, 1).value == 'hello world'
    assert db.wb['users'].cell(2, 1).value == 'id'
    assert db.wb['users'].cell(2, 2).value == 'name'
